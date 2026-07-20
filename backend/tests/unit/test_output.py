import csv
import io
import json

import pytest

from app.engines.output.base import ExportContext
from app.engines.output.registry import get_exporter, get_exporters

DATA = {
    "invoice_number": "INV-2026-0042", "invoice_date": "2026-03-15",
    "vendor_name": "ACME SUPPLIES LTD", "currency": "USD",
    "subtotal": 100.0, "tax_amount": 18.0, "discount": None, "total": 118.0,
    "line_items": [
        {"description": "Widget A", "quantity": 10.0, "unit_price": 5.0, "amount": 50.0},
        {"description": "Widget B", "quantity": 4.0, "unit_price": 12.5, "amount": 50.0},
    ],
}
CONF = {"overall": 1.0, "fields": [{"field": "total", "confidence": 1.0, "signals": ["present(+0.50)"]}]}


def _ctx(**options):
    return ExportContext(document_id="doc-1", filename="inv.pdf", data=DATA,
                         confidence=CONF, validation={"overall": "pass"}, options=options)


def test_all_exporters_autoload():
    assert set(get_exporters()) >= {"json", "csv", "xml", "xlsx"}


def test_unknown_format_raises_with_available_list():
    with pytest.raises(ValueError, match="Available"):
        get_exporter("pdf417")


def test_json_export_excludes_evidence_by_default():
    out = json.loads(get_exporter("json").render(_ctx()))
    assert out["data"]["total"] == 118.0
    assert "confidence" not in out


def test_json_export_includes_evidence_when_requested():
    out = json.loads(get_exporter("json").render(_ctx(include_evidence=True)))
    assert out["confidence"]["overall"] == 1.0
    assert out["validation"]["overall"] == "pass"


def test_csv_export_is_one_row_per_line_item():
    rows = list(csv.DictReader(io.StringIO(get_exporter("csv").render(_ctx()).decode())))
    assert len(rows) == 2
    assert rows[0]["invoice_number"] == "INV-2026-0042"
    assert rows[1]["line_description"] == "Widget B"


def test_csv_export_handles_zero_line_items():
    ctx = _ctx()
    ctx.data = {**DATA, "line_items": []}
    rows = list(csv.DictReader(io.StringIO(get_exporter("csv").render(ctx).decode())))
    assert len(rows) == 1
    assert rows[0]["line_amount"] == ""


def test_xml_export_is_wellformed():
    from xml.etree.ElementTree import fromstring
    root = fromstring(get_exporter("xml").render(_ctx()).decode())
    assert root.attrib["documentId"] == "doc-1"
    assert root.find("total").text == "118.0"
    assert len(root.find("lineItems")) == 2


def test_xlsx_export_produces_a_real_workbook():
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(get_exporter("xlsx").render(_ctx(include_evidence=True))))
    assert set(wb.sheetnames) == {"Invoice", "Line Items", "Evidence"}
    assert wb["Line Items"].max_row == 3


def test_filename_derives_from_source_document():
    assert get_exporter("csv").filename_for(_ctx()) == "inv.csv"
