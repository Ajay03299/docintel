import io

from openpyxl import Workbook

from app.engines.output.base import ExportContext, Exporter
from app.engines.output.registry import register_exporter


@register_exporter
class XlsxExporter(Exporter):
    format_id = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = ".xlsx"

    def render(self, ctx: ExportContext) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        ws.append(["Field", "Value"])
        for k, v in ctx.data.items():
            if k != "line_items":
                ws.append([k, v])

        items = wb.create_sheet("Line Items")
        items.append(["Description", "Quantity", "Unit Price", "Amount"])
        for item in ctx.data.get("line_items") or []:
            items.append([item.get("description"), item.get("quantity"),
                          item.get("unit_price"), item.get("amount")])

        if ctx.options.get("include_evidence", False):
            ev = wb.create_sheet("Evidence")
            ev.append(["Field", "Confidence", "Signals"])
            for f in ctx.confidence.get("fields", []):
                ev.append([f["field"], f["confidence"], "; ".join(f.get("signals", []))])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
